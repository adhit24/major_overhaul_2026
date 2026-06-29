"""
Migrasi data riil dari Workbook/induction-control-system.xlsm ke Supabase.

Script ini dijalankan SEKALI saja secara lokal (tidak pernah di-commit ke git,
tidak pernah jalan di Vercel). Data peserta berisi info pribadi karyawan,
jadi alurnya: baca file .xlsm lokal -> kirim langsung ke Supabase lewat REST
API pakai service_role key dari .env.local -> selesai, tidak ada file
perantara yang berisi data asli.

Cara pakai:
    cd major_overhaul_2026
    pip install openpyxl requests python-dotenv
    python scripts/migrate_data.py
"""

import datetime
import os

import openpyxl
import requests
from dotenv import load_dotenv

load_dotenv(".env.local")

SUPABASE_URL = os.environ["NEXT_PUBLIC_SUPABASE_URL"]
SERVICE_ROLE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
WB_PATH = r"..\SAFETY INDUCTION\Program Man Power Analyst\Workbook\induction-control-system.xlsm"

HEADERS = {
    "apikey": SERVICE_ROLE_KEY,
    "Authorization": f"Bearer {SERVICE_ROLE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal",
}

BULAN_ID = {
    "januari": 1, "februari": 2, "maret": 3, "april": 4, "mei": 5, "juni": 6,
    "juli": 7, "agustus": 8, "september": 9, "oktober": 10, "november": 11, "desember": 12,
}


def parse_tanggal(value):
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, str) and value.strip():
        parts = value.strip().lower().split()
        if len(parts) == 3 and parts[1] in BULAN_ID:
            try:
                return datetime.date(int(parts[2]), BULAN_ID[parts[1]], int(parts[0])).isoformat()
            except ValueError:
                return None
    return None


def yn_to_bool(value):
    if value is None:
        return None
    return str(value).strip().upper() == "Y"


def push(table, rows, batch_size=200):
    for i in range(0, len(rows), batch_size):
        chunk = rows[i : i + batch_size]
        resp = requests.post(f"{SUPABASE_URL}/rest/v1/{table}", headers=HEADERS, json=chunk, timeout=30)
        if resp.status_code >= 300:
            raise RuntimeError(f"Gagal insert ke {table} (baris {i}-{i+len(chunk)}): {resp.status_code} {resp.text}")
        print(f"  {table}: {i + len(chunk)}/{len(rows)} baris terkirim")


def migrate_peserta(wb):
    ws = wb["Database_Peserta"]
    rows = []
    for r in range(3, ws.max_row + 1):
        record_id = ws.cell(r, 1).value
        nama = ws.cell(r, 5).value
        if not record_id or not nama:
            continue
        rows.append(
            {
                "legacy_record_id": record_id,
                "no_badge": ws.cell(r, 2).value,
                "no_erp": str(ws.cell(r, 3).value) if ws.cell(r, 3).value is not None else None,
                "tanggal_induction": parse_tanggal(ws.cell(r, 4).value),
                "nama": str(nama).strip(),
                "job_no": ws.cell(r, 6).value,
                "departemen": ws.cell(r, 7).value,
                "kategori": ws.cell(r, 8).value,
                "jabatan_deskripsi": ws.cell(r, 9).value,
                "leader": ws.cell(r, 10).value,
                "ktp": yn_to_bool(ws.cell(r, 11).value),
                "sks": yn_to_bool(ws.cell(r, 12).value),
                "sertifikat": yn_to_bool(ws.cell(r, 13).value),
                "status_badge": ws.cell(r, 14).value or "PENDING",
                "due_date": parse_tanggal(ws.cell(r, 15).value),
                "remarks": ws.cell(r, 16).value,
            }
        )
    print(f"Database_Peserta: {len(rows)} baris siap dikirim")
    push("peserta", rows)


def migrate_deposit(wb):
    ws = wb["Summary_Deposit"]
    rows = []
    for r in range(3, ws.max_row + 1):
        no = ws.cell(r, 1).value
        tanggal = parse_tanggal(ws.cell(r, 2).value)
        if not no or not tanggal:
            continue
        rows.append(
            {
                "legacy_no": no,
                "tanggal": tanggal,
                "departemen_section": ws.cell(r, 3).value,
                "keterangan": ws.cell(r, 4).value,
                "rentang_no_id": str(ws.cell(r, 5).value) if ws.cell(r, 5).value is not None else None,
                "jumlah_kartu": ws.cell(r, 6).value,
                "tarif_kartu": ws.cell(r, 7).value or 50000,
                "due_date": parse_tanggal(ws.cell(r, 9).value),
                "status_batch": ws.cell(r, 10).value or "PENDING",
                "remarks": ws.cell(r, 11).value,
            }
        )
    print(f"Summary_Deposit: {len(rows)} baris siap dikirim")
    push("deposit_batch", rows)


def main():
    wb = openpyxl.load_workbook(WB_PATH, data_only=True)
    migrate_peserta(wb)
    migrate_deposit(wb)
    print("Selesai.")


if __name__ == "__main__":
    main()
